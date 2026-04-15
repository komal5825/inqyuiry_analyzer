import openpyxl


def fill_excel_template(template_path, output_path, data):
    """
    Fills the Infiniti PEB Specification Excel template with processed data.

    Cell mapping verified against actual template coordinates:
    ── PEB Specifications sheet ──────────────────────────────────────────────
    TABLE 1 (Metadata) — values go in column C, rows 2-10
      C2  = Date
      C3  = Structure Application (project)
      C4  = Proposal ID
      C6  = Building No.       (row 5 is blank)
      C7  = Option
      C8  = Revision
      C9  = Area sqm           (has formula =C14*C15, we overwrite with value)
      C10 = Design Software

    TABLE 2 (Geometry) — values go in column C, rows 13-30
      C13 = Building type
      C14 = Length in m
      C15 = Width in m
      C16 = Left eave height in m
      C17 = Right eave height in m
      C18 = Ridge line distance  (has formula =C15/2, we overwrite)
      C19 = Roof slope
      C20 = Left end wall bay spacing
      C21 = Right end wall bay spacing (has formula =C20, we overwrite)
      C22 = Side wall bay spacing
      C23 = Wall bracing type
      C24 = Roof bracing type
      C25 = Roof sheeting
      C26 = Wall sheeting
      C27 = Doors
      C28 = Windows
      C29 = Block wall in m

    CANOPY section — rows 32-34
      C32 = Canopy extension in m
      C33 = Canopy slope
      C34 = Clear top height in m

    ACCESSORIES — rows 36-39
      C36 = Braced bays
      C37 = Skylights
      C38 = Turbo ventilators
      C39 = Base elevation

    RIGHT-SIDE TABLE (Design params) — column H, rows 2-10
      H2  = Design code
      H3  = Dead load kN/m2
      H4  = Live load kN/m2
      H5  = Collateral load kN/m2
      H6  = Max wind speed m/s
      H7  = Max rainfall mm/hr
      H8  = Exposure category
      H9  = Seismic zone coefficient
      H10 = Max temperature variation

    DEFLECTIONS — column H, rows 13-18
      H13 = Frames vertical deflection
      H14 = Frames horizontal deflection
      H15 = Purlins
      H16 = Girts

    MATERIAL GRADES — column H, rows 21-31
      H21 = Hot rolled sections grade (MPa)
      H22 = Cold form sections grade (MPa)
      H23 = Primary sections grade (MPa)
      H25 = Min web thickness mm
      H26 = Min flange thickness mm
      H27 = Min flange width mm
      H28 = Min web depth mm
      H29 = Min cold form thickness mm
      H31 = Connection bolts grade

    ── Crane & mezzanine sheet ───────────────────────────────────────────────
      C2  = Crane capacity in kN
      C3  = Crane beam top height in m
      C4  = Crane span in m
      C5  = Crane running length in m
      C6  = Crane class
      C7  = Crane type
      C8  = Type of girder

      C18 = Mezzanine location
      C20 = Mezzanine live load
      C21 = Mezzanine dead load
      C22 = Mezzanine collateral load
      C24 = Mezzanine length in m
      C25 = Mezzanine width in m
      C26 = Top of mezzanine floor in m
      C27 = Depth of mezzanine slab in mm
    """
    wb = openpyxl.load_workbook(template_path)

    # ── PEB Specifications sheet ──────────────────────────────────────────────
    ws = wb['PEB Specifications']

    # TABLE 1 — Metadata
    ws['C2']  = data.get('date')
    ws['C3']  = data.get('project')
    ws['C4']  = data.get('proposal_id')
    ws['C6']  = data.get('building_no', 1)
    ws['C7']  = data.get('option', 1)
    ws['C8']  = data.get('revision', 0)
    ws['C9']  = data.get('area', 0)      # FIX: was formula =C14*C15; overwrite with computed value
    ws['C10'] = data.get('design_software', 'MBS')

    # TABLE 2 — Geometry
    ws['C13'] = data.get('building_type')
    ws['C14'] = data.get('length')
    ws['C15'] = data.get('width')
    ws['C16'] = data.get('left_height')   # FIX: was data.get('height') — now split left/right
    ws['C17'] = data.get('right_height')  # FIX: was also data.get('height') — duplicate
    ws['C18'] = data.get('ridge_dist')    # FIX: was formula =C15/2; overwrite with computed
    ws['C19'] = data.get('roof_slope', '10/100')
    ws['C20'] = data.get('end_bay')
    ws['C21'] = data.get('end_bay')       # Right end bay — same as left if symmetric
    ws['C22'] = data.get('side_bay')
    ws['C23'] = data.get('wall_bracing', 'Rod')
    ws['C24'] = data.get('roof_bracing', 'Rod')
    ws['C25'] = data.get('roof_sheeting', '0.45mm aluzinc')
    ws['C26'] = data.get('wall_sheeting', '0.45mm aluzinc')
    ws['C27'] = data.get('doors', '3x3')
    ws['C28'] = data.get('windows')       # FIX: was C28 missing entirely in original
    ws['C29'] = data.get('block_wall', 3.0)

    # CANOPY section
    ws['C32'] = data.get('canopy_ext', 0)
    ws['C33'] = data.get('roof_slope', '10/100')   # FIX: canopy slope (was missing)
    ws['C34'] = data.get('canopy_height', 0)

    # ACCESSORIES
    ws['C36'] = data.get('braced_bays', 0)
    ws['C37'] = data.get('skylights', 0)
    ws['C38'] = data.get('turbo_ventilators', 0)

    # RIGHT SIDE — Design parameters (column H)
    ws['H2']  = data.get('design_code', 'AISC')
    ws['H3']  = data.get('dead_load', 0.1)
    ws['H4']  = data.get('live_load', 0.57)
    ws['H5']  = data.get('collateral_load', 0.0)
    ws['H6']  = data.get('wind_speed', 39)
    ws['H7']  = data.get('rainfall', 150)
    ws['H8']  = data.get('exposure', 'B')
    ws['H9']  = data.get('seismic', 0.16)
    ws['H10'] = data.get('temp_var', 20)

    # DEFLECTIONS (column H)
    ws['H13'] = data.get('v_deflection', 'Span/150')
    ws['H14'] = data.get('h_deflection', 'Height/100')
    ws['H15'] = data.get('purlin_defl',  'Span/150')   # FIX: was missing in original
    ws['H16'] = data.get('girt_defl',    'Span/120')   # FIX: was missing in original

    # MATERIAL GRADES (column H)
    ws['H21'] = data.get('hr_grade', 250)
    ws['H22'] = data.get('cf_grade', 350)
    ws['H23'] = data.get('primary_grade', 350)
    ws['H25'] = data.get('min_web_thickness', 4)
    ws['H26'] = data.get('min_flange_thickness', 5)
    ws['H27'] = data.get('min_flange_width', 125)
    ws['H28'] = data.get('min_web_depth', 250)
    ws['H31'] = data.get('connection_bolts', 'High strength bolts (ASTM A325 or equivalent)')

    # ── Crane & mezzanine sheet ───────────────────────────────────────────────
    ws2 = wb['Crane & mezzanine']

    # Crane
    ws2['C2'] = data.get('crane_capacity') or None
    ws2['C3'] = data.get('crane_top_ht') or None
    ws2['C4'] = None   # crane_span — not in logic_engine output yet; placeholder
    ws2['C5'] = None   # crane_running_length — placeholder
    ws2['C6'] = data.get('crane_class', 'II (Normal duty)')
    ws2['C7'] = data.get('crane_type', 'Top running')
    ws2['C8'] = data.get('girder_type', 'Single')

    # Mezzanine
    ws2['C18'] = None  # location — not computed yet
    ws2['C20'] = None  # mezzanine live load
    ws2['C21'] = None  # mezzanine dead load
    ws2['C22'] = None  # mezzanine collateral load
    ws2['C24'] = data.get('mezzanine_length') or None
    ws2['C25'] = data.get('mezzanine_width') or None
    ws2['C26'] = data.get('mezzanine_height') or None

    wb.save(output_path)
    return output_path